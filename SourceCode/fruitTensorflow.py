import tensorflow as tf
from openpyxl import load_workbook

# read excel file
wb      = load_workbook('감_분석.xlsx',read_only=True)
data    = wb["분석데이터"]
print("DATA LOAD")

# 분석할 달
month   = 1

# 우리의 Y 값
marketPrice     = list()

# X values
importWeight    = list()
r1_f            = list()
r2_f            = list()
r3_f            = list()
r4_f            = list()
r5_f            = list()
r6_f            = list()
r7_f            = list()
r8_f            = list()
r9_f            = list()
r10_f           = list()
r11_f           = list()
r12_f           = list()
r13_f           = list()
r14_f           = list()
r15_f           = list()
r16_f           = list()
r17_f           = list()
r1_t            = list()
r1_r            = list()
r1_s            = list()
r2_t            = list()
r2_r            = list()
r2_s            = list()
r3_t            = list()
r3_r            = list()
r3_s            = list()
r4_t            = list()
r4_r            = list()
r4_s            = list()
r5_t            = list()
r5_r            = list()
r5_s            = list()
r6_t            = list()
r6_r            = list()
r6_s            = list()
r7_t            = list()
r7_r            = list()
r7_s            = list()
r8_t            = list()
r8_r            = list()
r8_s            = list()
r9_t            = list()
r9_r            = list()
r9_s            = list()
r10_t           = list()
r10_r           = list()
r10_s           = list()
r11_t           = list()
r11_r           = list()
r11_s           = list()
r12_t           = list()
r12_r           = list()
r12_s           = list()
r13_t           = list()
r13_r           = list()
r13_s           = list()
r14_t           = list()
r14_r           = list()
r14_s           = list()
r15_t           = list()
r15_r           = list()
r15_s           = list()
r16_t           = list()
r16_r           = list()
r16_s           = list()
r17_t           = list()
r17_r           = list()
r17_s           = list()
xfeed = list()            

# 각 변수에 모으기
for row in data.iter_rows():
    if (row[0].value == month and row[1].value != 0):
        marketPrice.append(row[1].value)
        importWeight.append(row[2].value)
        r1_f.append(row[3].value)
        r2_f.append(row[4].value)
        r3_f.append(row[5].value)
        r4_f.append(row[6].value)
        r5_f.append(row[7].value)
        r6_f.append(row[8].value)
        r7_f.append(row[9].value)
        r8_f.append(row[10].value)
        r9_f.append(row[11].value)
        r10_f.append(row[12].value)
        r11_f.append(row[13].value)
        r12_f.append(row[14].value)
        r13_f.append(row[15].value)
        r14_f.append(row[16].value)
        r15_f.append(row[17].value)
        r16_f.append(row[18].value)
        r17_f.append(row[19].value)
        r1_t.append(row[20].value)
        r1_r.append(row[21].value)
        r1_s.append(row[22].value)
        r2_t.append(row[23].value)
        r2_r.append(row[24].value)
        r2_s.append(row[25].value)
        r3_t.append(row[26].value)
        r3_r.append(row[27].value)
        r3_s.append(row[28].value)
        r4_t.append(row[29].value)
        r4_r.append(row[30].value)
        r4_s.append(row[31].value)
        r5_t.append(row[32].value)
        r5_r.append(row[33].value)
        r5_s.append(row[34].value)
        r6_t.append(row[35].value)
        r6_r.append(row[36].value)
        r6_s.append(row[37].value)
        r7_t.append(row[38].value)
        r7_r.append(row[39].value)
        r7_s.append(row[40].value)
        r8_t.append(row[41].value)
        r8_r.append(row[42].value)
        r8_s.append(row[43].value)
        r9_t.append(row[44].value)
        r9_r.append(row[45].value)
        r9_s.append(row[46].value)
        r10_t.append(row[47].value)
        r10_r.append(row[48].value)
        r10_s.append(row[49].value)
        r11_t.append(row[50].value)
        r11_r.append(row[51].value)
        r11_s.append(row[52].value)
        r12_t.append(row[53].value)
        r12_r.append(row[54].value)
        r12_s.append(row[55].value)
        r13_t.append(row[56].value)
        r13_r.append(row[57].value)
        r13_s.append(row[58].value)
        r14_t.append(row[59].value)
        r14_r.append(row[60].value)
        r14_s.append(row[61].value)
        r15_t.append(row[62].value)
        r15_r.append(row[63].value)
        r15_s.append(row[64].value)
        r16_t.append(row[65].value)
        r16_r.append(row[66].value)
        r16_s.append(row[67].value)
        r17_t.append(row[68].value)
        r17_r.append(row[69].value)
        r17_s.append(row[70].value)
        xfeed.append([row[2].value, row[15].value, row[56].value, row[57].value, row[58].value])


print("DATA INIT")

# reqularizer setup
reqularizer = tf.contrib.l2_reqularizer(scale = 0.00001)

# Y values placeholder
y_marketPrice     = tf.placeholder(tf.float32)

# X values placeholder
xs = tf.placeholder(tf.float32)

w_weight = tf.get_variable("weights",[5,220],reqularizer=reqularizer, initializer=tf.contrib.layer.xavier_initializer())

b = tf.Variable(tf.random_normal([1]))
    
print("Tensorflow INIT")

hypothesis = tf.matmul(xs,w_weights) + b
cost = tf.reduce_mean(tf.square(hypothesis - y_marketPrice))
optimizer = tf.train.GradientDescentOptimizer(learning_rate=1e-10)
train = optimizer.minimize(cost)

reg_variables = tf.get_collection(tf.GraphKeys.REQULARIZATION_LOSSES)
reg_term = tf.contrib.layers.apply_reqularization(reqularizer,reg_variables)
cost += reg_term

sess = tf.Session()
sess.run(tf.global_variables_initializer())

for step in range (2001):
    cost_val, hy_val, _ = sess.run ( [cost, hypothesis, train],
        feed_dict = {  xs:xfeed, y_marketPrice:marketPrice } )
    if step % 10 == 0:
        print(step, "Cost: ", cost_val, "\nPrediction:\n", hy_val)